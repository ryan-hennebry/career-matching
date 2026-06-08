#!/usr/bin/env python3
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(HERE, "output", "_sheet_rows.json")
OUT = os.path.join(HERE, "output", "The Search - Jobs - Ranked.xlsx")

with open(SRC) as f:
    data = json.load(f)

COLUMNS = [
    "Firm", "Title", "Location", "Category", "Posted", "Experience", "Salary",
    "Tokens", "Agentic", "Applied", "Messaged", "Response", "Chased",
    "Interview", "Declined",
]

# ---- styles ----
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FILL = PatternFill("solid", fgColor="EDEDED")
BANNER_FILL = PatternFill("solid", fgColor="F3F3F3")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(bold=True, color="000000")
BANNER_FONT = Font(bold=True, color="000000")
PLAIN_FONT = Font(color="000000")
LINK_FONT = Font(color="1155CC", underline="single")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

LOCATION_COLORS = {
    "London": "8AA1B1",
    "Remote": "0B6E99",
    "Cambridge": "0B6E99",
}
CATEGORY_COLORS = {
    "Applied AI": "3D6B9E",
    "Early-stage": "137333",
    "Operations": "1A73E8",
    "Product": "7F7F7F",
    "Marketing": "D2691E",
}
CHECK_COLS = ["Agentic", "Applied", "Messaged", "Response", "Chased", "Interview", "Declined"]
EMPTY_COLS = ["Posted", "Experience", "Salary", "Tokens"]

def glyph(v):
    return "☑" if str(v).upper() == "TRUE" else "☐"

wb = Workbook()

# ================= SHEET 1: Jobs =================
ws = wb.active
ws.title = "Jobs"

# header
for ci, name in enumerate(COLUMNS, start=1):
    c = ws.cell(row=1, column=ci, value=name)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER
ws.freeze_panes = "A2"

banner_count = 0
r = 2
for item in data["rows"]:
    if "banner" in item and len(item) == 1:
        banner_count += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        c = ws.cell(row=r, column=1, value=item["banner"])
        c.fill = BANNER_FILL
        c.font = BANNER_FONT
        c.alignment = CENTER
        # apply fill+border across merged range
        for ci in range(1, 16):
            cc = ws.cell(row=r, column=ci)
            cc.fill = BANNER_FILL
            cc.border = BORDER
        r += 1
        continue

    for ci, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=r, column=ci)
        c.border = BORDER
        c.fill = WHITE_FILL
        if name == "Firm":
            c.value = item.get("Firm", "")
            c.font = PLAIN_FONT
            c.alignment = LEFT
        elif name == "Title":
            title = item.get("Title", "")
            url = item.get("title_url", "") or ""
            if url:
                c.value = title
                c.hyperlink = url
                c.font = LINK_FONT
            else:
                c.value = title
                c.font = PLAIN_FONT
            c.alignment = LEFT
        elif name == "Location":
            loc = item.get("Location", "")
            c.value = loc
            color = LOCATION_COLORS.get(loc, "434343")
            c.font = Font(color=color)
            c.alignment = LEFT
        elif name == "Category":
            cat = item.get("Category", "")
            c.value = cat
            color = CATEGORY_COLORS.get(cat, "434343")
            c.font = Font(color=color)
            c.alignment = LEFT
        elif name in EMPTY_COLS:
            c.value = ""
            c.alignment = CENTER
        elif name in CHECK_COLS:
            c.value = glyph(item.get(name, "FALSE"))
            c.font = PLAIN_FONT
            c.alignment = CENTER
    r += 1

# column widths
widths = {
    "Firm": 22, "Title": 46, "Location": 16, "Category": 14,
    "Posted": 10, "Experience": 10, "Salary": 10, "Tokens": 10,
    "Agentic": 10, "Applied": 10, "Messaged": 10, "Response": 10,
    "Chased": 10, "Interview": 10, "Declined": 10,
}
from openpyxl.utils import get_column_letter
for ci, name in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = widths[name]

sheet1_total = r - 1  # rows written excluding header? -> header is row1, last data row = r-1

# ================= SHEET 2: Excluded =================
ws2 = wb.create_sheet("Excluded")
EXC_COLS = ["Firm", "Title", "Reason", "Link"]
for ci, name in enumerate(EXC_COLS, start=1):
    c = ws2.cell(row=1, column=ci, value=name)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

er = 2
for item in data["excluded"]:
    url = item.get("title_url", "") or ""
    # Firm
    c = ws2.cell(row=er, column=1, value=item.get("Firm", ""))
    c.border = BORDER; c.fill = WHITE_FILL; c.font = PLAIN_FONT; c.alignment = LEFT
    # Title
    c = ws2.cell(row=er, column=2, value=item.get("Title", ""))
    c.border = BORDER; c.fill = WHITE_FILL; c.alignment = LEFT
    if url:
        c.hyperlink = url
        c.font = LINK_FONT
    else:
        c.font = PLAIN_FONT
    # Reason
    c = ws2.cell(row=er, column=3, value=item.get("reason", ""))
    c.border = BORDER; c.fill = WHITE_FILL; c.font = PLAIN_FONT; c.alignment = LEFT
    # Link
    c = ws2.cell(row=er, column=4, value=url)
    c.border = BORDER; c.fill = WHITE_FILL; c.font = PLAIN_FONT; c.alignment = LEFT
    er += 1

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 46
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 60

wb.save(OUT)

# ---- verify ----
from openpyxl import load_workbook
wb2 = load_workbook(OUT)
s1 = wb2["Jobs"]
s2 = wb2["Excluded"]
# count banners in sheet1 by scanning merged-cell first column values matching banner strings
banners = {b["banner"] for b in data["rows"] if "banner" in b and len(b) == 1}
bc = 0
for row in s1.iter_rows(min_row=2, max_col=1):
    v = row[0].value
    if v in banners:
        bc += 1
print("PATH:", OUT)
print("Sheet1 'Jobs' total rows (incl header):", s1.max_row)
print("Sheet1 data rows (excl header):", s1.max_row - 1)
print("Sheet1 banner rows:", bc)
print("Sheet2 'Excluded' total rows (incl header):", s2.max_row)
print("Sheet2 data rows (excl header):", s2.max_row - 1)
